import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.app import models
from backend.app.services.normalization import (
    MEETING_TYPES,
    REGIONS,
    normalize_key,
    normalize_name,
    normalize_text,
    parse_bool,
    parse_datetime,
    parse_int,
    split_multi,
)
from backend.app.services.skill_catalog import canonical_skill, catalog_skill


RESERVED_IMPORT_FIELDS = {
    "external_id",
    "full_name",
    "region",
    "country",
    "business_unit",
    "role_title",
    "seniority_level",
    "languages",
    "profile_summary",
    "summary",
    "skills",
    "industries",
    "bandwidth",
    "validation_status",
    "manager_note",
    "last_updated_at",
    "project_name",
    "role_on_project",
    "project_status",
    "meeting_comfort",
    "comfort_note",
    "intro_call_comfort",
    "executive_pitch_comfort",
    "workshop_comfort",
    "deep_dive_comfort",
    "client_facing_desire",
    "travel_preference",
    "stretch_interest",
    "timestamp",
    "start_time",
    "completion_time",
    "email",
    "email_address",
    "last_modified_time",
    "industry_experience_notes",
    "client_facing_experiences",
    "client_facing_engagement_notes",
}

REGION_ALIASES = {
    "europe": "EMEA",
    "emea": "EMEA",
    "north america": "NA",
    "na": "NA",
    "usa": "NA",
    "us": "NA",
    "united states": "NA",
    "canada": "NA",
    "uk": "UK",
    "united kingdom": "UK",
    "hong kong": "HK",
    "hk": "HK",
    "malaysia": "MY",
    "my": "MY",
    "singapore": "SG",
    "sg": "SG",
    "australia": "AUS",
    "aus": "AUS",
    "austria": "EMEA",
    "belgium": "EMEA",
    "france": "EMEA",
    "germany": "EMEA",
    "ireland": "EMEA",
    "italy": "EMEA",
    "luxembourg": "EMEA",
    "netherlands": "EMEA",
    "poland": "EMEA",
    "south africa": "EMEA",
    "spain": "EMEA",
    "switzerland": "EMEA",
    "uae": "EMEA",
    "united arab emirates": "EMEA",
    "new zealand": "AUS",
}


class ImportRowProblem(Exception):
    def __init__(self, field: str, message: str) -> None:
        super().__init__(message)
        self.field = field
        self.message = message


def normalize_region_value(value: object) -> str:
    text = normalize_text(value)
    return REGION_ALIASES.get(text.lower(), text)


def normalize_header(value: object) -> str:
    header = normalize_key(str(value))
    if header in RESERVED_IMPORT_FIELDS:
        return header
    if "email" in header:
        return "email"
    if "start" in header and "time" in header:
        return "start_time"
    if "completion" in header and "time" in header:
        return "completion_time"
    if "modified" in header and "time" in header:
        return "last_modified_time"
    if "region" in header or "market" in header:
        return "region"
    if "industry" in header and "notes" in header:
        return "industry_experience_notes"
    if "industry" in header and ("sector" in header or "experience" in header):
        return "industries"
    if "client" in header and "facing" in header and "notes" in header:
        return "client_facing_engagement_notes"
    if "client" in header and "facing" in header and "experience" in header:
        return "client_facing_experiences"
    if "appetite" in header and "client" in header:
        return "client_facing_desire"
    if "country" in header or "location" in header or "office" in header or "geo" in header:
        return "country"
    name_like_headers = {
        "trainer",
        "coach",
        "consultant",
        "resource",
        "employee",
        "trainer_name",
        "coach_name",
        "consultant_name",
        "resource_name",
        "employee_name",
        "which_skills_lab_coach_are_you",
    }
    if (
        header in name_like_headers
        or header.startswith("which_skills_lab_coach_are_you")
        or (header.endswith("_name") and "manager" not in header)
    ):
        return "full_name"
    return header


def parse_skill_level(value: object) -> int | None:
    text = normalize_text(value).lower()
    if not text or text in {"0", "n", "no", "none", "n/a", "na", "not applicable", "no experience"}:
        return None
    if text in {"1", "2", "3", "4", "5"}:
        return parse_int(text)
    if text in {"yes", "y", "true", "x", "interested", "open"}:
        return 3
    if text in {"basic", "beginner", "aware", "awareness", "limited experience"}:
        return 2
    if text in {"intermediate", "working", "can support", "competent"}:
        return 3
    if text in {"advanced", "strong", "experienced"}:
        return 4
    if text in {"expert", "lead", "can lead", "sme"}:
        return 5
    return None


def parse_skills(value: object) -> list[dict[str, Any]]:
    items = split_multi(value)
    skills: list[dict[str, Any]] = []
    for item in items:
        name = item
        level = 3
        if ":" in item:
            name, rating = item.rsplit(":", 1)
            level = parse_int(rating)
        skill_name, skill_category = canonical_skill(name)
        skills.append(
            {
                "skill_name": skill_name,
                "skill_category": skill_category,
                "proficiency_level": level,
                "evidence_type": "Import",
                "evidence_note": "Imported from trainer spreadsheet",
            }
        )
    return [skill for skill in skills if skill["skill_name"]]


def capability_skills_from_record(record: dict[str, Any]) -> list[dict[str, Any]]:
    skills: list[dict[str, Any]] = []
    for key, value in record.items():
        if key in RESERVED_IMPORT_FIELDS:
            continue
        if key.startswith("additional_skills"):
            for skill in parse_skills(value):
                skill["skill_category"] = "Additional Skills"
                skills.append(skill)
            continue
        if key.endswith("_note") or "comment" in key or "explain" in key:
            continue
        level = parse_skill_level(value)
        if level is None:
            continue
        matched_skill = catalog_skill(key)
        if matched_skill is None:
            continue
        skill_name, skill_category = matched_skill
        skills.append(
            {
                "skill_name": skill_name,
                "skill_category": skill_category,
                "proficiency_level": level,
                "evidence_type": "Import",
                "evidence_note": f"Imported from {skill_category} capability column",
            }
        )
    return skills


def format_skill_name(key: str) -> str:
    return canonical_skill(key)[0]


def skills_from_record(record: dict[str, Any]) -> list[dict[str, Any]]:
    explicit = parse_skills(record.get("skills"))
    return explicit if explicit else capability_skills_from_record(record)


def score_header_row(row: list[Any]) -> int:
    headers = [normalize_header(header) for header in row]
    non_empty = [header for header in headers if header and header != "none"]
    if len(non_empty) < 2:
        return 0

    score = 0
    if "full_name" in headers:
        score += 20
    if "region" in headers:
        score += 10
    if "country" in headers:
        score += 8
    if "email" in headers:
        score += 4
    if "start_time" in headers or "completion_time" in headers:
        score += 3
    score += sum(2 for header in headers if header in RESERVED_IMPORT_FIELDS)
    score += sum(1 for header in headers if header and header not in RESERVED_IMPORT_FIELDS)

    # A lone title like "FDM Skills Lab Coach Profiling Form" should not win.
    return score if score >= 20 else 0


def find_header_index(rows: list[list[Any]]) -> int | None:
    best_index: int | None = None
    best_score = 0
    for index, row in enumerate(rows[:20]):
        score = score_header_row(row)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def normalize_rows(rows: list[list[Any]]) -> list[tuple[int, dict[str, Any]]]:
    if not rows:
        return []
    header_index = find_header_index(rows)
    if header_index is None:
        return []
    headers = [normalize_header(header) for header in rows[header_index]]
    normalized: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        record = {}
        for index, header in enumerate(headers):
            if header:
                record[header] = row[index] if index < len(row) else ""
        if any(normalize_text(value) for value in record.values()):
            normalized.append((row_number, record))
    return normalized


def preview_headers(rows: list[list[Any]]) -> str:
    previews: list[str] = []
    for row in rows[:10]:
        cells = [normalize_text(cell) for cell in row if normalize_text(cell)]
        if cells:
            previews.append(", ".join(cells[:8]))
        if len(previews) >= 3:
            break
    return " | ".join(previews) if previews else "No non-empty cells found"


async def read_upload_rows(file: UploadFile) -> tuple[str, list[list[Any]]]:
    suffix = Path(file.filename or "").suffix.lower()
    content = await file.read()
    if suffix == ".csv":
        decoded = content.decode("utf-8-sig")
        return "csv", list(csv.reader(StringIO(decoded)))
    if suffix == ".xlsx":
        # Some Microsoft Forms exports have a stale sheet dimension of A1 even
        # though the table spans many columns. Normal mode reads the real cells.
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
        best_rows: list[list[Any]] = []
        best_score = 0
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            header_index = find_header_index(sheet_rows)
            score = score_header_row(sheet_rows[header_index]) if header_index is not None else 0
            if score > best_score:
                best_rows = sheet_rows
                best_score = score
        return "xlsx", best_rows
    if suffix == ".xls":
        raise ImportRowProblem("file", "Legacy .xls is not supported yet. Save as .xlsx or .csv.")
    raise ImportRowProblem("file", "Unsupported file type. Upload .xlsx or .csv.")


def validate_record(record: dict[str, Any]) -> None:
    if not normalize_text(record.get("full_name")):
        raise ImportRowProblem("full_name", "Trainer name is required")
    region = normalize_region_value(record.get("region") or record.get("country"))
    if not region:
        raise ImportRowProblem("region", "Region is required")
    if region not in REGIONS:
        raise ImportRowProblem("region", f"Unsupported region. Use one of: {', '.join(sorted(REGIONS))}")
    if not skills_from_record(record):
        raise ImportRowProblem("skills", "At least one skill or rated capability column is required")


def find_existing_trainer(session: Session, record: dict[str, Any]) -> models.Trainer | None:
    external_id = normalize_text(record.get("external_id"))
    if external_id:
        trainer = session.scalar(select(models.Trainer).where(models.Trainer.external_id == external_id))
        if trainer:
            return trainer

    return session.scalar(
        select(models.Trainer).where(
            models.Trainer.full_name_normalized == normalize_name(record["full_name"]),
            models.Trainer.region == normalize_region_value(record.get("region") or record.get("country")),
        )
    )


def upsert_trainer(session: Session, record: dict[str, Any]) -> bool:
    trainer = find_existing_trainer(session, record)
    created = trainer is None
    if trainer is None:
        trainer = models.Trainer(
            external_id=normalize_text(record.get("external_id")) or None,
            full_name=normalize_text(record["full_name"]),
            full_name_normalized=normalize_name(record["full_name"]),
            region=normalize_region_value(record.get("region") or record.get("country")),
        )
        session.add(trainer)

    trainer.external_id = normalize_text(record.get("external_id")) or trainer.external_id
    trainer.full_name = normalize_text(record["full_name"])
    trainer.full_name_normalized = normalize_name(record["full_name"])
    trainer.region = normalize_region_value(record.get("region") or record.get("country"))
    trainer.country = normalize_text(record.get("country")) or None
    trainer.business_unit = normalize_text(record.get("business_unit")) or None
    trainer.role_title = normalize_text(record.get("role_title")) or "Trainer"
    trainer.seniority_level = normalize_text(record.get("seniority_level")) or "Senior"
    trainer.languages = "; ".join(split_multi(record.get("languages")) or ["English"])
    trainer.profile_summary = normalize_text(record.get("profile_summary") or record.get("summary")) or None
    trainer.bandwidth = normalize_text(record.get("bandwidth")) or "Medium"
    trainer.validation_status = normalize_text(record.get("validation_status")) or "Self-declared"
    trainer.manager_note = normalize_text(record.get("manager_note")) or None
    trainer.last_updated_at = parse_datetime(record.get("last_updated_at"))

    trainer.skills.clear()
    imported_skills = skills_from_record(record)
    for skill in imported_skills:
        trainer.skills.append(models.Skill(**skill))

    industries = split_multi(record.get("industries"))
    project_name = normalize_text(record.get("project_name")) or "Imported profile context"
    trainer.projects.clear()
    trainer.projects.append(
        models.Project(
            project_name=project_name,
            client_sector=industries[0] if industries else None,
            role_on_project=normalize_text(record.get("role_on_project")) or None,
            project_status=normalize_text(record.get("project_status")) or "Recent",
            time_commitment=trainer.bandwidth,
            relevance_tags="; ".join([skill["skill_name"] for skill in imported_skills]),
        )
    )

    trainer.comfort.clear()
    if not created:
        session.flush()
    default_comfort = parse_int(record.get("meeting_comfort"), default=3)
    for meeting_type in MEETING_TYPES:
        field_key = f"{meeting_type.lower().replace(' ', '_')}_comfort"
        trainer.comfort.append(
            models.MeetingComfort(
                meeting_type=meeting_type,
                comfort_level=parse_int(record.get(field_key), default=default_comfort),
                confidence_note=normalize_text(record.get("comfort_note")) or "Imported profile needs comfort detail.",
                validated_by_manager=trainer.validation_status != "Self-declared",
            )
        )

    if trainer.preference is None:
        trainer.preference = models.Preference()
    trainer.preference.client_facing_desire = normalize_text(record.get("client_facing_desire")) or "Neutral"
    trainer.preference.travel_preference = normalize_text(record.get("travel_preference")) or "Regional"
    trainer.preference.preferred_sectors = "; ".join(industries)
    trainer.preference.stretch_interest = parse_bool(record.get("stretch_interest"))

    return created


async def import_trainers(session: Session, file: UploadFile) -> models.ImportBatch:
    source_type, rows = await read_upload_rows(file)
    records = normalize_rows(rows)
    batch = models.ImportBatch(filename=file.filename or "upload", source_type=source_type)
    session.add(batch)
    session.flush()

    if not records:
        batch.error_count = 1
        batch.errors.append(
            models.ImportRowError(
                row_number=1,
                field="file",
                message=(
                    "No importable trainer rows found. The sheet must include a trainer name column and either "
                    f"region, country, or location. First rows seen: {preview_headers(rows)}"
                ),
            )
        )

    for row_number, record in records:
        try:
            validate_record(record)
            created = upsert_trainer(session, record)
            if created:
                batch.created_count += 1
            else:
                batch.updated_count += 1
        except ImportRowProblem as exc:
            batch.error_count += 1
            batch.errors.append(models.ImportRowError(row_number=row_number, field=exc.field, message=exc.message))

    session.commit()
    session.refresh(batch)
    return batch
