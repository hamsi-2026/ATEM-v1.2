from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import delete

from backend.app import models
from backend.app.database import SessionLocal
from backend.app.main import app


TEST_TRAINER_NAMES = {
    "Test Trainer",
    "Capability Column Trainer",
    "Form Header Trainer",
    "Second Sheet Trainer",
    "Coverage Duplicate Trainer",
}


@pytest.fixture(autouse=True)
def cleanup_test_trainers() -> None:
    cleanup_generated_trainers()
    yield
    cleanup_generated_trainers()


def cleanup_generated_trainers() -> None:
    with SessionLocal() as session:
        trainer_ids = [
            trainer_id
            for (trainer_id,) in session.query(models.Trainer.id)
            .filter(models.Trainer.full_name.in_(TEST_TRAINER_NAMES))
            .all()
        ]
        if trainer_ids:
            session.execute(delete(models.MatchResult).where(models.MatchResult.trainer_id.in_(trainer_ids)))
            session.execute(delete(models.Skill).where(models.Skill.trainer_id.in_(trainer_ids)))
            session.execute(delete(models.Project).where(models.Project.trainer_id.in_(trainer_ids)))
            session.execute(delete(models.MeetingComfort).where(models.MeetingComfort.trainer_id.in_(trainer_ids)))
            session.execute(delete(models.Preference).where(models.Preference.trainer_id.in_(trainer_ids)))
            session.execute(delete(models.Trainer).where(models.Trainer.id.in_(trainer_ids)))
        session.commit()


def test_health() -> None:
    with TestClient(app) as client:
        assert client.get("/health").json() == {"status": "ok"}


def test_skill_catalog_comes_from_form_categories() -> None:
    with TestClient(app) as client:
        response = client.get("/config/skill-catalog")

    assert response.status_code == 200
    catalog = response.json()
    categories = {group["category"]: group["skills"] for group in catalog}
    assert "Software Engineering" in categories
    assert "AI/ML" in categories
    assert "Data Analytics/Engineering" in categories
    assert "AI Governance" in categories["AI/ML"]
    assert "Data Visualization With Power BI" in categories["Data Analytics/Engineering"]


def test_xlsx_import_creates_trainer() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "region", "country", "languages", "skills", "industries", "bandwidth"])
    sheet.append(["Test Trainer", "SG", "Singapore", "English", "AI governance:5", "Retail", "Low"])
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    with TestClient(app) as client:
        response = client.post(
            "/imports/trainers",
            files={"file": ("trainers.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    body = response.json()
    assert body["error_count"] == 0
    assert body["created_count"] + body["updated_count"] >= 1


def test_xlsx_import_finds_later_header_and_capability_columns() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["FDM Skills Lab Coach Profiling Form"])
    sheet.append([])
    sheet.append(["coach name", "region", "country", "AI governance", "Cloud modernisation"])
    sheet.append(["Capability Column Trainer", "United Kingdom", "United Kingdom", "Expert", 4])
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    with TestClient(app) as client:
        response = client.post(
            "/imports/trainers",
            files={"file": ("capabilities.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    body = response.json()
    assert body["error_count"] == 0
    assert body["created_count"] + body["updated_count"] >= 1
    with SessionLocal() as session:
        trainer = session.query(models.Trainer).filter_by(full_name="Capability Column Trainer").one()
        skills = {skill.skill_name: skill for skill in trainer.skills}
        assert skills["AI Governance"].skill_category == "AI/ML"
        assert skills["AI Governance"].proficiency_level == 5
        assert skills["Cloud Modernisation"].skill_category == "Cloud"
        assert skills["Cloud Modernisation"].proficiency_level == 4


def test_xlsx_import_accepts_trainer_detail_form_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["FDM Skills Lab Coach Profiling Form"])
    sheet.append(["Start time", "Email", "Which Skills Lab Coach are you?", "Your work location", "Cybersecurity"])
    sheet.append(["2026-05-24", "coach@example.com", "Form Header Trainer", "Singapore", "Advanced"])
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    with TestClient(app) as client:
        response = client.post(
            "/imports/trainers",
            files={"file": ("form_headers.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    body = response.json()
    assert body["error_count"] == 0
    assert body["created_count"] + body["updated_count"] >= 1
    with SessionLocal() as session:
        trainer = session.query(models.Trainer).filter_by(full_name="Form Header Trainer").one()
        skill = trainer.skills[0]
        assert skill.skill_name == "Cybersecurity"
        assert skill.skill_category == "Testing"
        assert skill.proficiency_level == 4


def test_xlsx_import_uses_form_categories_and_excludes_client_appetite() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "Which Skills Lab Coach are you?",
        "Region",
        "Appetite for Client Work (1 = Low Interest, 5 = High Interest)",
        "Data Visualization with Power Bi",
        "Random Excel Metric",
        "Additional Skills - Are there any skills not listed above that you believe are relevant to your role?",
    ])
    sheet.append(["Capability Column Trainer", "SG", 5, "Competent", 5, "oil and gas analytics:4"])
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    with TestClient(app) as client:
        response = client.post(
            "/imports/trainers",
            files={"file": ("categories.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    body = response.json()
    assert body["error_count"] == 0
    with SessionLocal() as session:
        trainer = session.query(models.Trainer).filter_by(full_name="Capability Column Trainer").one()
        skills = {skill.skill_name: skill for skill in trainer.skills}
        assert "Appetite For Client Work" not in skills
        assert "Random Excel Metric" not in skills
        assert skills["Data Visualization With Power BI"].skill_category == "Data Analytics/Engineering"
        assert skills["Data Visualization With Power BI"].proficiency_level == 3
        assert skills["Oil And Gas Analytics"].skill_category == "Additional Skills"


def test_xlsx_import_scans_all_sheets_for_trainer_details() -> None:
    workbook = Workbook()
    workbook.active.title = "Instructions"
    workbook.active.append(["FDM Skills Lab Coach Profiling Form"])
    sheet = workbook.create_sheet("Responses")
    sheet.append(["Start time", "Resource name", "Office location", "Data analytics"])
    sheet.append(["2026-05-24", "Second Sheet Trainer", "Hong Kong", 5])
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    with TestClient(app) as client:
        response = client.post(
            "/imports/trainers",
            files={"file": ("second_sheet.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    body = response.json()
    assert body["error_count"] == 0
    assert body["created_count"] + body["updated_count"] >= 1


def test_coverage_counts_unique_trainers_not_duplicate_skill_rows() -> None:
    with SessionLocal() as session:
        trainer = models.Trainer(
            full_name="Coverage Duplicate Trainer",
            full_name_normalized="coverage duplicate trainer",
            region="SG",
            country="Singapore",
            role_title="Trainer",
            seniority_level="Senior",
            languages="English",
        )
        trainer.skills.append(
            models.Skill(skill_name="Coverage Unique Skill", skill_category="Additional Skills", proficiency_level=4)
        )
        trainer.skills.append(
            models.Skill(skill_name="Coverage Unique Skill", skill_category="Additional Skills", proficiency_level=5)
        )
        session.add(trainer)
        session.commit()

    with TestClient(app) as client:
        response = client.get("/analytics/coverage")

    assert response.status_code == 200
    row = next(row for row in response.json()["rows"] if row["topic"] == "Coverage Unique Skill")
    cell = next(cell for cell in row["cells"] if cell["region"] == "SG")
    assert cell["best"] == 5
    assert cell["count"] == 1
    assert cell["trainer_names"] == ["Coverage Duplicate Trainer"]
